# ============================================================================
# src/index_generator.py
# ============================================================================
#
# Genera y actualiza el index.html principal del proyecto.
#
# Uso desde cualquier notebook:
#   from src.index_generator import actualizar_index
#   actualizar_index()
#
# La función lee dinámicamente los datos disponibles (Excel, parquets)
# y regenera el index.html con KPIs actualizados.
# Si un dato no está disponible (fase no ejecutada), el KPI aparece
# translúcido con "Pendiente".
#
# NO tiene parámetros — todo se importa de src.config.
# ============================================================================

from datetime import datetime
from pathlib import Path


def actualizar_index(verbose=True):
    """
    Regenera docs/html/index.html con KPIs dinámicos.

    Lee los datos disponibles (Excel originales, parquets de cada fase)
    y construye el HTML con los KPIs actualizados.

    Parameters
    ----------
    verbose : bool
        Si True, imprime progreso por consola.

    Returns
    -------
    Path
        Ruta del index.html generado.
    """
    import pandas as pd

    # --- Imports de config (todo centralizado) ---
    from src.config import (
        RUTA_RAW, RUTA_FEATURES, RUTA_HTML,
        EXCEL_PRINCIPAL, EXCEL_PREINSCRIPCION,
        DATASET_FINAL_PARQUET,
        AUTORA, EMAIL_UOC, EMAIL_UJI, GITHUB_REPO,
        COLORES,
    )

    if verbose:
        print('🔄 Actualizando index.html...')

    # =================================================================
    # PASO 1: Datos del Excel con openpyxl (rápido, read_only)
    # =================================================================
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PRINCIPAL, read_only=True)
        hojas = wb.sheetnames
        n_tablas = len(hojas)
        n_registros_raw = 0

        for hoja in hojas:
            ws = wb[hoja]
            n_filas = ws.max_row - 1 if ws.max_row else 0
            n_registros_raw += n_filas
        wb.close()

        if EXCEL_PREINSCRIPCION.exists():
            wb2 = load_workbook(EXCEL_PREINSCRIPCION, read_only=True)
            ws2 = wb2[wb2.sheetnames[0]]
            n_registros_raw += ws2.max_row - 1 if ws2.max_row else 0
            n_tablas += 1
            wb2.close()
    except Exception as e:
        if verbose:
            print(f'  ⚠ Error leyendo Excel: {e}')
        hojas = []
        n_tablas = 9
        n_registros_raw = 0

    n_archivos_excel = len(list(RUTA_RAW.glob('*.xlsx')))

    # =================================================================
    # PASO 2: KPIs dinámicos de parquets (solo si existen)
    # =================================================================
    kpi_expedientes = None
    kpi_periodo = None
    kpi_n_anios = None
    kpi_variables = None
    kpi_abandono = None

    # Fase 1: df_alumno.parquet
    if DATASET_FINAL_PARQUET.exists():
        df = pd.read_parquet(DATASET_FINAL_PARQUET)
        for col in ['curso_aca', 'curso_aca_id']:
            if col in df.columns:
                p_min = int(df[col].min())
                p_max = int(df[col].max())
                kpi_periodo = f'{p_min}-{p_max}'
                kpi_n_anios = p_max - p_min + 1
                break
        for col in ['per_id_ficticio', 'nip', 'alumno_id']:
            if col in df.columns:
                kpi_expedientes = df[col].nunique()
                break
        del df

    # Fase 3: features
    ruta_feat = RUTA_FEATURES / 'df_expediente_features.parquet'
    if ruta_feat.exists():
        df_feat = pd.read_parquet(ruta_feat)
        kpi_variables = len(df_feat.columns) - 1
        if 'abandono' in df_feat.columns:
            kpi_abandono = round((df_feat['abandono'] == 1).mean() * 100, 1)
        del df_feat

    # =================================================================
    # PASO 3: Construir HTML
    # =================================================================
    fecha = datetime.now().strftime('%d/%m/%Y')

    def kpi_html(valor, label, color, disponible=True):
        opacity = '' if disponible else ' opacity:0.35;'
        texto = valor if disponible else 'Pendiente'
        return (
            f'\n                <div class="kpi" style="{opacity}">'
            f'\n                    <div class="value" style="color:{color}">{texto}</div>'
            f'\n                    <div class="label">{label}</div>'
            f'\n                </div>'
        )

    kpis_gen = ''
    kpis_gen += kpi_html(
        f'{kpi_expedientes:,}'.replace(',', '.') if kpi_expedientes else '',
        'Expedientes únicos', COLORES.PRINCIPAL, kpi_expedientes is not None)
    kpis_gen += kpi_html(
        f'{kpi_periodo} ({kpi_n_anios} años)' if kpi_periodo else '',
        'Período', COLORES.OK, kpi_periodo is not None)
    kpis_gen += kpi_html(
        f'{kpi_variables}+' if kpi_variables else '',
        'Variables finales', COLORES.ALERTA, kpi_variables is not None)
    kpis_gen += kpi_html(
        f'~{kpi_abandono:.0f}%' if kpi_abandono else '',
        'Tasa abandono', COLORES.ERROR, kpi_abandono is not None)

    hay_pendientes = (kpi_expedientes is None or kpi_variables is None)
    aviso = ''
    if hay_pendientes:
        aviso = (
            '\n            <div style="background:#EBF8FF; padding:15px; border-radius:8px;'
            ' margin-top:15px; border-left:4px solid #3182ce; opacity:0.8;">'
            '\n                <strong>ℹ️ Nota:</strong> Los KPIs translúcidos se activarán'
            ' cuando ejecutes las fases correspondientes y vuelvas a ejecutar este notebook.'
            '\n            </div>'
        )

    hojas_texto = ', '.join(hojas + ['Preinscripción']) if hojas else 'No disponible'
    assets_rel = '../assets'

    # --- HTML completo ---
    H = []
    H.append('<!DOCTYPE html>')
    H.append('<html lang="es">')
    H.append('<head>')
    H.append('    <meta charset="UTF-8">')
    H.append('    <meta name="viewport" content="width=device-width, initial-scale=1.0">')
    H.append('    <title>TFM: Predicción de Abandono Universitario</title>')
    H.append('    <link rel="stylesheet" href="style.css">')
    H.append('</head>')
    H.append('<body>')
    H.append('    <header>')
    H.append('        <div class="header-content">')
    H.append(f'            <img src="{assets_rel}/logo_uoc.png" alt="UOC" class="header-logo" onerror="this.style.display=\'none\'">')
    H.append('            <div class="header-title">')
    H.append('                <h1>🎓 TFM: Predicción de Abandono Universitario</h1>')
    H.append('                <p>Análisis y Modelado de Datos de Estudiantes UJI</p>')
    H.append('            </div>')
    H.append(f'            <img src="{assets_rel}/logo_uji.jpg" alt="UJI" class="header-logo" onerror="this.style.display=\'none\'">')
    H.append('        </div>')
    H.append('    </header>')
    H.append('')
    H.append('    <nav class="nav-fases">')
    H.append('        <a href="index.html" class="active">🏠 Inicio</a>')
    H.append('        <a href="fase1/fase1_index.html">📥 Transformación</a>')
    H.append('        <a href="fase2/fase2_index.html">📊 EDA Raw</a>')
    H.append('        <a href="fase3/fase3_index.html">🔧 Features</a>')
    H.append('        <a href="fase4/fase4_index.html">🔬 EDA Final</a>')
    H.append('        <a href="fase5/fase5_index.html" class="disabled">🤖 Modelado</a>')
    H.append('        <a href="fase6/fase6_index.html" class="disabled">📈 Evaluación</a>')
    H.append('        <a href="fase7/fase7_index.html" class="disabled">🚀 Aplicación</a>')
    H.append('        <a href="fase_automl/fase_automl_index.html">⚡ Pre-Modelado: AutoML</a>')
    H.append('    </nav>')
    H.append('')
    H.append('    <div class="container">')
    H.append('')
    H.append('        <section class="seccion">')
    H.append('            <h2>📋 Estructura del Proyecto</h2>')
    H.append('            <p>Haz clic en una fase para explorar todos sus módulos.</p>')
    H.append('            <div class="grid-tarjetas">')
    H.append('                <a href="fase1/fase1_index.html" class="tarjeta" style="border-left-color:#3182ce">')
    H.append('                    <h3 style="color:#3182ce">📥 Fase 1: Transformación</h3>')
    H.append(f'                    <p>Limpieza, unificación de {n_tablas} tablas y creación de df_alumno.</p>')
    H.append('                    <div class="link" style="color:#3182ce">Abrir fase →</div>')
    H.append('                </a>')
    H.append('                <a href="fase2/fase2_index.html" class="tarjeta" style="border-left-color:#38a169">')
    H.append('                    <h3 style="color:#38a169">📊 Fase 2: EDA Raw</h3>')
    H.append('                    <p>Análisis exploratorio de los datos originales.</p>')
    H.append('                    <div class="link" style="color:#38a169">Abrir fase →</div>')
    H.append('                </a>')
    H.append('                <a href="fase3/fase3_index.html" class="tarjeta" style="border-left-color:#805ad5">')
    H.append('                    <h3 style="color:#805ad5">🔧 Fase 3: Feature Engineering</h3>')
    H.append('                    <p>Ingeniería de variables y preparación para modelado.</p>')
    H.append('                    <div class="link" style="color:#805ad5">Abrir fase →</div>')
    H.append('                </a>')
    H.append('                <a href="fase4/fase4_index.html" class="tarjeta" style="border-left-color:#ed8936">')
    H.append('                    <h3 style="color:#ed8936">🔬 Fase 4: EDA Final</h3>')
    H.append('                    <p>Análisis exploratorio del dataset final de features.</p>')
    H.append('                    <div class="link" style="color:#ed8936">Abrir fase →</div>')
    H.append('                </a>')
    H.append('                <a href="fase5/fase5_index.html" class="tarjeta" style="border-left-color:#e53e3e; opacity:0.6">')
    H.append('                    <h3 style="color:#e53e3e">🤖 Fase 5: Modelado</h3>')
    H.append('                    <p>Entrenamiento de modelos de clasificación supervisada.</p>')
    H.append('                    <div class="link" style="color:#e53e3e">Próximamente</div>')
    H.append('                </a>')
    H.append('                <a href="fase6/fase6_index.html" class="tarjeta" style="border-left-color:#319795; opacity:0.6">')
    H.append('                    <h3 style="color:#319795">📈 Fase 6: Evaluación</h3>')
    H.append('                    <p>Comparativa de modelos, métricas e interpretabilidad.</p>')
    H.append('                    <div class="link" style="color:#319795">Próximamente</div>')
    H.append('                </a>')
    H.append('                <a href="fase7/fase7_index.html" class="tarjeta" style="border-left-color:#2c5282; opacity:0.6">')
    H.append('                    <h3 style="color:#2c5282">🚀 Fase 7: Aplicación</h3>')
    H.append('                    <p>Dashboard, vistas por perfil y documentación final.</p>')
    H.append('                    <div class="link" style="color:#2c5282">Próximamente</div>')
    H.append('                </a>')
    H.append('            </div>')
    H.append('        </section>')
    H.append('')
    H.append('        <section class="seccion">')
    H.append('            <h2>⚡ Pre-Modelado: AutoML</h2>')
    H.append('            <div style="background:#FFFBEB; padding:20px; border-radius:10px; border-left:4px solid #D69E2E; margin-bottom:15px;">')
    H.append('                <strong>⚠️ Requisito:</strong> Para usar AutoML, primero ejecuta las Fases 1, 2 y 3.')
    H.append('            </div>')
    H.append('            <a href="fase_automl/fase_automl_index.html"')
    H.append('               style="display:inline-block; padding:12px 24px; background:#5B21B6; color:white;')
    H.append('                      border-radius:8px; text-decoration:none; font-weight:bold;">')
    H.append('                ⚡ Exploración de Frameworks AutoML →')
    H.append('            </a>')
    H.append('        </section>')
    H.append('')
    H.append('        <section class="seccion">')
    H.append('            <h2>📊 Datos del Proyecto</h2>')
    H.append('            <div style="background:#f8fafc; padding:20px; border-radius:10px; border-left:4px solid #3182ce; margin-bottom:20px;">')
    H.append(f'                <p><strong>📁 Origen:</strong> {n_archivos_excel} archivos Excel con {n_tablas} tablas</p>')
    H.append(f'                <p><strong>📋 Tablas:</strong> {hojas_texto}</p>')
    H.append(f'                <p><strong>📊 Registros totales (raw):</strong> {n_registros_raw:,}</p>')
    H.append('            </div>')
    H.append('        </section>')
    H.append('')
    H.append('        <section class="seccion">')
    H.append('            <h2>📈 Estadísticas del Proyecto</h2>')
    H.append(f'            <div class="kpis">{kpis_gen}')
    H.append(f'            </div>{aviso}')
    H.append('        </section>')
    H.append('')
    H.append('    </div>')
    H.append('')
    H.append('    <footer>')
    H.append(f'        <strong>{AUTORA}</strong> |')
    H.append(f'        <a href="mailto:{EMAIL_UOC}">{EMAIL_UOC}</a> (UOC) |')
    H.append(f'        <a href="mailto:{EMAIL_UJI}">{EMAIL_UJI}</a> (UJI)')
    H.append('        <br>')
    H.append(f'        <a href="{GITHUB_REPO}" target="_blank">📓 GitHub</a>')
    H.append(f'        <br>Generado: {fecha}')
    H.append('    </footer>')
    H.append('</body>')
    H.append('</html>')

    html = '\n'.join(H)

    # --- Guardar ---
    ruta_index = RUTA_HTML / 'index.html'
    ruta_index.write_text(html, encoding='utf-8')

    if verbose:
        estado_kpis = []
        if kpi_expedientes:
            estado_kpis.append(f'Expedientes: {kpi_expedientes:,}'.replace(',', '.'))
        if kpi_periodo:
            estado_kpis.append(f'Período: {kpi_periodo}')
        if kpi_variables:
            estado_kpis.append(f'Variables: {kpi_variables}+')
        if kpi_abandono:
            estado_kpis.append(f'Abandono: ~{kpi_abandono:.0f}%')

        if estado_kpis:
            print(f'  ✅ KPIs: {" | ".join(estado_kpis)}')
        else:
            print(f'  ⚠ KPIs pendientes (ejecuta las fases primero)')
        print(f'  ✅ {ruta_index}')

    return ruta_index
